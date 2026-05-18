"""
Report generation: Excel (.xlsx), plain-text, and optional Obsidian markdown.

Encryption: set REPORTS_ENCRYPTION_KEY in your .env to enable.
  - Excel files are password-protected with that key.
  - Text files are Fernet-encrypted (saved as .txt.enc, decryptable via /api/reports/<id>/text).

Rotation: files older than REPORTS_RETENTION_DAYS (default 365) are deleted
automatically at the start of each report generation.
"""
import hashlib
import io
import logging
import os
import re
from base64 import urlsafe_b64encode
from datetime import datetime, timezone, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Built-in defaults when /obsidian is mounted (override via env only if needed)
_DEFAULT_OBSIDIAN_VAULT = "/obsidian"
_DEFAULT_REPORTS_SUBDIR = (
    "10_Personal/13_Finances/Investments/AI Investment Analysis"
)
_DEFAULT_KNOWLEDGE_SUBDIR = "50_Knowledge/notes"

_COLOURS = {
    "BUY":        "2196F3",
    "SELL":       "F44336",
    "HOLD_LONG":  "4CAF50",
    "HOLD_SHORT": "FF9800",
    "HOLD":       "9E9E9E",
    "WATCH":      "9C27B0",
    "POSITIVE":   "4CAF50",
    "NEUTRAL":    "9E9E9E",
    "NEGATIVE":   "F44336",
}

# Badge text colour — dark text on light badges
_BADGE_TEXT = {
    "HOLD_SHORT": "000000",
    "WATCH":      "FFFFFF",
}


def _derive_fernet_key(password: str) -> bytes:
    """Derive a valid 32-byte Fernet key from an arbitrary password string."""
    digest = hashlib.sha256(password.encode()).digest()
    return urlsafe_b64encode(digest)


class ReportGenerator:
    def __init__(self, reports_dir: Optional[str] = None,
                 encryption_key: Optional[str] = None,
                 retention_days: int = 365,
                 obsidian_vault_dir: Optional[str] = None,
                 obsidian_reports_subdir: Optional[str] = None,
                 obsidian_knowledge_subdir: Optional[str] = None,
                 obsidian_knowledge_moc_dir: Optional[str] = None,
                 obsidian_knowledge_enabled: Optional[bool] = None,
                 obsidian_default_moc: Optional[str] = None):
        self.reports_dir = reports_dir or os.getenv("REPORTS_DIR", "/data/reports")
        self.retention_days = int(os.getenv("REPORTS_RETENTION_DAYS", str(retention_days)))
        raw_key = encryption_key or os.getenv("REPORTS_ENCRYPTION_KEY", "")
        self._enc_key: Optional[str] = raw_key if raw_key else None
        os.makedirs(self.reports_dir, exist_ok=True)

        vault = obsidian_vault_dir if obsidian_vault_dir is not None else None
        if vault is None:
            vault = os.getenv("OBSIDIAN_VAULT_DIR", "").strip()
        if not vault and os.path.isdir(_DEFAULT_OBSIDIAN_VAULT):
            vault = _DEFAULT_OBSIDIAN_VAULT
        self.obsidian_vault_dir = vault if vault else None
        self._obsidian_vault_auto = (
            not (os.getenv("OBSIDIAN_VAULT_DIR") or "").strip()
            and self.obsidian_vault_dir == _DEFAULT_OBSIDIAN_VAULT
        )

        if obsidian_reports_subdir is not None:
            sub = obsidian_reports_subdir
        else:
            sub = os.getenv("OBSIDIAN_REPORTS_SUBDIR", _DEFAULT_REPORTS_SUBDIR)
        self.obsidian_reports_subdir = sub.strip().strip("/\\") if sub else ""

        if obsidian_knowledge_subdir is not None:
            kn_sub = obsidian_knowledge_subdir
        else:
            kn_sub = os.getenv("OBSIDIAN_KNOWLEDGE_SUBDIR", _DEFAULT_KNOWLEDGE_SUBDIR)
        self.obsidian_knowledge_subdir = kn_sub.strip().strip("/\\") if kn_sub else ""

        moc_dir = (obsidian_knowledge_moc_dir if obsidian_knowledge_moc_dir is not None
                   else os.getenv("OBSIDIAN_KNOWLEDGE_MOC_DIR", "50_Knowledge/_moc"))
        self.obsidian_knowledge_moc_dir = moc_dir.strip().strip("/\\") if moc_dir else ""

        if obsidian_knowledge_enabled is not None:
            self.obsidian_knowledge_enabled = obsidian_knowledge_enabled
        else:
            raw = os.getenv("OBSIDIAN_KNOWLEDGE_ENABLED", "true").strip().lower()
            self.obsidian_knowledge_enabled = raw not in ("0", "false", "no", "off")

        default_moc = (obsidian_default_moc if obsidian_default_moc is not None
                       else os.getenv("OBSIDIAN_DEFAULT_MOC", "MOC-investment-analysis"))
        self.obsidian_default_moc = (default_moc or "").strip()

        self.reports_full_subdir = _env_subdir(
            "REPORTS_FULL_SUBDIR", "full",
        )
        self.reports_single_subdir = _env_subdir(
            "REPORTS_SINGLE_SUBDIR", "single",
        )
        self.obsidian_full_subdir = _env_subdir(
            "OBSIDIAN_REPORTS_FULL_SUBDIR", "Full Portfolio",
        )
        self.obsidian_single_subdir = _env_subdir(
            "OBSIDIAN_REPORTS_SINGLE_SUBDIR", "Individual Stock",
        )

    def _run_scope_key(self, run_scope: str) -> str:
        return "single" if (run_scope or "").strip().lower() == "single" else "full"

    def _disk_reports_dir(self, run_scope: str) -> str:
        """/data/reports/full or /data/reports/single"""
        scope = self._run_scope_key(run_scope)
        sub = (
            self.reports_single_subdir if scope == "single"
            else self.reports_full_subdir
        )
        path = os.path.join(self.reports_dir, sub)
        os.makedirs(path, exist_ok=True)
        return path

    def _obsidian_reports_dir(self, run_scope: str) -> str:
        """Vault path: …/AI Investment Analysis/Full Portfolio|Individual Stock"""
        scope = self._run_scope_key(run_scope)
        sub = (
            self.obsidian_single_subdir if scope == "single"
            else self.obsidian_full_subdir
        )
        base = self.obsidian_reports_subdir
        path = os.path.join(self.obsidian_vault_dir, base, sub) if base else ""
        if path:
            os.makedirs(path, exist_ok=True)
        return path

    def _moc_runs_section(self, run_scope: str) -> str:
        return (
            "Individual Stock runs"
            if self._run_scope_key(run_scope) == "single"
            else "Full Portfolio runs"
        )

    def obsidian_moc_active(self) -> bool:
        return bool(
            self.obsidian_vault_dir
            and self.obsidian_knowledge_moc_dir
            and self.obsidian_default_moc
        )

    def obsidian_knowledge_active(self) -> bool:
        return (
            self.obsidian_knowledge_enabled
            and self.obsidian_vault_dir
            and self.obsidian_knowledge_subdir
        )

    # ── Excel ──────────────────────────────────────────────────────────────────

    def generate_excel(
        self, run_id: int, analyses: list[dict], *, run_scope: str = "full",
    ) -> str:
        self._rotate_old_reports()

        filename = f"analysis_run{run_id}_{_date_stamp()}.xlsx"
        path = os.path.join(self._disk_reports_dir(run_scope), filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis"

        headers = [
            "Ticker", "Rec", "Confidence", "Current Price", "Cost Basis",
            "P&L %", "P&L £", "Shares", "30d Target", "Target Lo", "Target Hi",
            "P/E", "EPS Growth %", "Analyst Target", "Analyst Consensus",
            "News Sentiment", "Reasoning", "Catalysts", "Risks", "Worries",
            "90d Outlook",
        ]
        _write_header_row(ws, headers)

        for row_num, a in enumerate(analyses, start=2):
            price  = a.get("current_price") or 0
            cost   = a.get("cost_basis") or 0
            shares = a.get("shares") or 0
            pnl_pct = ((price - cost) / cost * 100) if cost else 0
            pnl_abs = (price - cost) * shares

            values = [
                a.get("ticker", ""),
                a.get("recommendation", ""),
                a.get("confidence", ""),
                price,
                cost,
                pnl_pct / 100,
                pnl_abs,
                shares,
                a.get("price_target_30d"),
                a.get("price_target_lo"),
                a.get("price_target_hi"),
                a.get("pe_ratio"),
                a.get("eps_growth_pct"),
                a.get("analyst_target_mean"),
                a.get("analyst_consensus", "").upper() if a.get("analyst_consensus") else "",
                a.get("news_sentiment", ""),
                a.get("reasoning", ""),
                "\n".join(a.get("catalysts", [])),
                "\n".join(a.get("risks", [])),
                "\n".join(a.get("worries", [])),
                a.get("outlook_90d", ""),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            rec = a.get("recommendation", "")
            if rec in _COLOURS:
                ws.cell(row=row_num, column=2).fill = PatternFill(
                    fill_type="solid", fgColor=_COLOURS[rec]
                )
                text_colour = _BADGE_TEXT.get(rec, "FFFFFF")
                ws.cell(row=row_num, column=2).font = Font(color=text_colour, bold=True)

            ws.cell(row=row_num, column=6).number_format = "0.00%"

        col_widths = [10, 12, 12, 14, 12, 10, 10, 10, 12, 10, 10,
                      8, 12, 14, 16, 16, 50, 40, 40, 30, 50]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "C2"
        wb.save(path)

        if self._enc_key:
            path = self._encrypt_excel(path)

        logger.info("Excel report saved to %s", path)
        return path

    # ── Plain text ─────────────────────────────────────────────────────────────

    def generate_text(
        self, run_id: int, analyses: list[dict], *, run_scope: str = "full",
    ) -> str:
        self._rotate_old_reports()

        ext = "txt.enc" if self._enc_key else "txt"
        filename = f"analysis_run{run_id}_{_date_stamp()}.{ext}"
        path = os.path.join(self._disk_reports_dir(run_scope), filename)

        lines = [
            f"STOCK ANALYSIS REPORT — Run #{run_id}",
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            "=" * 72,
        ]
        for a in analyses:
            price  = a.get("current_price") or 0
            cost   = a.get("cost_basis") or 0
            shares = a.get("shares") or 0
            pnl_pct = ((price - cost) / cost * 100) if cost else 0
            lines += [
                "",
                "─" * 72,
                f"  {a.get('ticker', '?')}  [{a.get('recommendation', '?')}]  "
                f"Confidence: {a.get('confidence', '?')}",
                "─" * 72,
                f"  Price: {price:.2f}  Cost: {cost:.2f}  P&L: {pnl_pct:+.1f}%  Shares: {shares}",
                f"  30d Target: {a.get('price_target_30d', '?')}  "
                f"Range: {a.get('price_target_lo', '?')} – {a.get('price_target_hi', '?')}",
                "",
            ]
            if a.get("pe_ratio"):
                lines.append(f"  P/E: {a['pe_ratio']}  "
                             f"EPS growth: {a.get('eps_growth_pct', '?')}%  "
                             f"Analyst target: {a.get('analyst_target_mean', '?')} "
                             f"({a.get('analyst_consensus', '').upper()})")
                lines.append("")
            lines += [
                f"  Reasoning: {a.get('reasoning', '')}",
                "",
                f"  News ({a.get('news_sentiment', '?')}): {a.get('news_summary', '')}",
                "",
                "  Catalysts:",
                *[f"    + {c}" for c in a.get("catalysts", [])],
                "  Risks:",
                *[f"    - {r}" for r in a.get("risks", [])],
                "  Worries:",
                *[f"    ! {w}" for w in a.get("worries", [])],
                "",
                f"  90d Outlook: {a.get('outlook_90d', '')}",
            ]
        lines.append("")
        content = "\n".join(lines).encode("utf-8")

        if self._enc_key:
            content = self._fernet().encrypt(content)
            with open(path, "wb") as f:
                f.write(content)
        else:
            with open(path, "w", encoding="utf-8") as f:
                f.write(content.decode("utf-8"))

        logger.info("Text report saved to %s", path)
        return path

    # ── Obsidian markdown (knowledge base) ─────────────────────────────────────

    def obsidian_enabled(self) -> bool:
        return bool(self.obsidian_vault_dir and self.obsidian_reports_subdir)

    def obsidian_status(self) -> dict:
        """Diagnostics for vault mount, env config, and recent .md reports."""
        vault = self.obsidian_vault_dir or ""
        base = self.obsidian_reports_subdir or ""
        full_dir = (
            os.path.join(vault, base, self.obsidian_full_subdir)
            if vault and base else ""
        )
        single_dir = (
            os.path.join(vault, base, self.obsidian_single_subdir)
            if vault and base else ""
        )

        reasons: list[str] = []
        if not vault:
            reasons.append("OBSIDIAN_VAULT_DIR is not set (set to /obsidian)")
        elif not os.path.isdir(vault):
            reasons.append(f"Vault path does not exist in container: {vault}")
        if not base:
            reasons.append("OBSIDIAN_REPORTS_SUBDIR is not set")

        mount_ok = bool(vault and os.path.isdir(vault))
        writable = False
        write_error: Optional[str] = None
        if mount_ok:
            probe = os.path.join(vault, ".stock-analyser-write-test")
            try:
                with open(probe, "w", encoding="utf-8") as f:
                    f.write("ok")
                os.remove(probe)
                writable = True
            except OSError as exc:
                write_error = str(exc)
                reasons.append(f"Vault not writable: {exc}")

        def _list_md(folder: str, limit: int = 8) -> list[str]:
            if not folder or not os.path.isdir(folder):
                return []
            try:
                names = [
                    n for n in os.listdir(folder)
                    if n.endswith(".md") and os.path.isfile(os.path.join(folder, n))
                ]
                names.sort(reverse=True)
                return names[:limit]
            except OSError:
                return []

        return {
            "enabled": self.obsidian_enabled(),
            "vault_auto_detected": getattr(self, "_obsidian_vault_auto", False),
            "vault_dir": vault or None,
            "reports_base": base or None,
            "full_portfolio_dir": full_dir or None,
            "individual_stock_dir": single_dir or None,
            "mount_exists": mount_ok,
            "vault_writable": writable,
            "write_error": write_error,
            "knowledge_enabled": self.obsidian_knowledge_active(),
            "default_moc": self.obsidian_default_moc or None,
            "recent_full": _list_md(full_dir),
            "recent_single": _list_md(single_dir),
            "issues": reasons,
            "ready": self.obsidian_enabled() and mount_ok and writable,
        }

    def obsidian_skip_reason(self) -> Optional[str]:
        if not self.obsidian_vault_dir:
            return (
                f"No vault at {_DEFAULT_OBSIDIAN_VAULT} — add Obsidian path mapping "
                "in Docker (host vault → container /obsidian)"
            )
        if not os.path.isdir(self.obsidian_vault_dir):
            return f"Vault path missing in container: {self.obsidian_vault_dir}"
        if not self.obsidian_reports_subdir:
            return "OBSIDIAN_REPORTS_SUBDIR is empty"
        return None

    def generate_markdown(
        self, run_id: int, analyses: list[dict], *, run_scope: str = "full",
    ) -> Optional[str]:
        """
        Write an Obsidian-friendly .md report into the vault (not rotated/deleted).
        Requires OBSIDIAN_VAULT_DIR + OBSIDIAN_REPORTS_SUBDIR (or constructor args).
        """
        skip = self.obsidian_skip_reason()
        if skip:
            logger.warning("Obsidian markdown skipped: %s", skip)
            return None

        dest_dir = self._obsidian_reports_dir(run_scope)
        if not dest_dir:
            return None

        day = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        scope_key = self._run_scope_key(run_scope)
        if scope_key == "single" and len(analyses) == 1:
            ticker = analyses[0].get("ticker", "position")
            filename = f"{day} {ticker} Analysis Run {run_id}.md"
        else:
            filename = f"{day} Analysis Run {run_id}.md"
        path = os.path.join(dest_dir, filename)

        tickers = sorted({a.get("ticker", "") for a in analyses if a.get("ticker")})
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        lines = [
            "---",
            f"run_id: {run_id}",
            f"scope: {scope_key}",
            f"generated: {generated}",
            "tags:",
            "  - investment-analysis",
            "  - stock-analyzer",
            "tickers:",
            *[f"  - {t}" for t in tickers],
        ]
        if self.obsidian_default_moc:
            lines.append(f"moc: '[[{self.obsidian_default_moc}]]'")
        lines += [
            "---",
            "",
            f"# Stock analysis — run {run_id}",
            "",
            f"Generated {generated} (UTC). Scope: **{scope_key}**.",
            "",
        ]

        for a in analyses:
            ticker = a.get("ticker", "?")
            rec = a.get("recommendation", "?")
            conf = a.get("confidence", "?")
            price = a.get("current_price") or 0
            cost = a.get("cost_basis") or 0
            shares = a.get("shares") or 0
            pnl_pct = ((price - cost) / cost * 100) if cost else 0

            lines += [
                f"## {ticker}",
                "",
                f"**{rec}** · {conf} confidence",
                "",
                "| | |",
                "|---|---|",
                f"| Price | {price:.4f} |" if price else "| Price | — |",
                f"| Cost | {cost:.4f} |" if cost else "| Cost | — |",
                f"| Shares | {shares:g} |" if shares else "| Shares | — |",
                f"| P&L | {pnl_pct:+.1f}% |" if cost else "| P&L | — |",
                f"| 30d target | {a.get('price_target_30d', '—')} |",
                "",
            ]
            if a.get("pe_ratio") or a.get("analyst_target_mean"):
                lines.append(
                    f"P/E {a.get('pe_ratio', '—')} · EPS growth {a.get('eps_growth_pct', '—')}% · "
                    f"Analyst target {a.get('analyst_target_mean', '—')} "
                    f"({(a.get('analyst_consensus') or '').upper()})"
                )
                lines.append("")
            if a.get("reasoning"):
                lines += ["### Thesis", "", str(a["reasoning"]), ""]
            if a.get("news_summary") or a.get("news_sentiment"):
                lines += [
                    "### News",
                    "",
                    f"**{a.get('news_sentiment', '?')}** — {a.get('news_summary', '')}",
                    "",
                ]
            cats = a.get("catalysts") or []
            if cats:
                lines += ["### Catalysts", ""]
                lines += [f"- {c}" for c in cats]
                lines.append("")
            risks = a.get("risks") or []
            if risks:
                lines += ["### Risks", ""]
                lines += [f"- {r}" for r in risks]
                lines.append("")
            worries = a.get("worries") or []
            if worries:
                lines += ["### Worries", ""]
                lines += [f"- {w}" for w in worries]
                lines.append("")
            if a.get("outlook_90d"):
                lines += ["### 90-day outlook", "", str(a["outlook_90d"]), ""]
            lines.append("---")
            lines.append("")

        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines).rstrip() + "\n")

        if self.obsidian_moc_active():
            report_wikilink = filename.removesuffix(".md")
            self._link_in_moc(
                self.obsidian_default_moc,
                report_wikilink,
                section=self._moc_runs_section(run_scope),
            )

        logger.info("Obsidian report saved to %s", path)
        return path

    def write_knowledge_notes(
        self,
        run_id: int,
        ticker: str,
        notes: list[dict],
        *,
        market_ticker: Optional[str] = None,
        run_report_day: Optional[str] = None,
    ) -> list[str]:
        """
        Write atomic notes to 50_Knowledge/notes when Claude returns knowledge_notes.
        Updates linked MOC files under 50_Knowledge/_moc when requested.
        """
        if not self.obsidian_knowledge_active() or not notes:
            return []

        notes_dir = os.path.join(self.obsidian_vault_dir, self.obsidian_knowledge_subdir)
        os.makedirs(notes_dir, exist_ok=True)

        day = run_report_day or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        run_link = f"[[{day} Analysis Run {run_id}]]"
        written: list[str] = []

        for item in notes:
            if not isinstance(item, dict):
                continue
            slug = _sanitize_knowledge_slug(item.get("slug") or "")
            body = (item.get("body") or item.get("summary") or "").strip()
            if not slug or not body:
                continue

            mocs = self._mocs_for_note(item)

            existing = _find_note_by_slug(notes_dir, slug)
            if existing:
                filename = existing
                note_id = filename.removesuffix(".md")
            else:
                note_id = f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{slug}"
                filename = f"{note_id}.md"

            path = os.path.join(notes_dir, filename)
            generated = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            title = (item.get("title") or slug.replace("-", " ")).strip()

            lines = [
                "---",
                "source: stock-analyzer",
                f"run_id: {run_id}",
                f"ticker: {ticker}",
            ]
            if market_ticker and market_ticker.upper() != ticker.upper():
                lines.append(f"market_ticker: {market_ticker}")
            lines += [
                f"created: {generated}",
                "tags:",
                "  - investment-knowledge",
                "  - stock-analyzer",
            ]
            if mocs:
                lines.append("mocs:")
                lines += [f"  - {m}" for m in mocs]
            lines += [
                "---",
                "",
                f"# {title}",
                "",
                body,
                "",
                "## Links",
                "",
                f"- Portfolio run: {run_link}",
                f"- Position: {ticker}",
            ]
            if mocs:
                lines.append("- MOCs: " + ", ".join(f"[[{m}]]" for m in mocs))

            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines).rstrip() + "\n")

            written.append(path)
            logger.info("Knowledge note %s → %s", "updated" if existing else "created", path)

            for moc_name in mocs:
                self._link_in_moc(moc_name, note_id, section="Knowledge notes")

        return written

    def _mocs_for_note(self, item: dict) -> list[str]:
        """Always include default MOC; merge Claude-suggested topic MOCs."""
        extra = item.get("mocs") or item.get("link_mocs") or []
        if isinstance(extra, str):
            extra = [extra]
        names: list[str] = []
        seen: set[str] = set()

        def add(name: str) -> None:
            name = (name or "").strip()
            if not name or name in seen:
                return
            seen.add(name)
            names.append(name)

        if self.obsidian_default_moc:
            add(self.obsidian_default_moc)
        for m in extra:
            add(str(m))
        return names

    def _ensure_moc_file(self, moc_name: str) -> str:
        """Create MOC file if missing; return path."""
        moc_file = moc_name if moc_name.endswith(".md") else f"{moc_name}.md"
        moc_path = os.path.join(
            self.obsidian_vault_dir, self.obsidian_knowledge_moc_dir, moc_file,
        )
        if os.path.isfile(moc_path):
            return moc_path

        os.makedirs(os.path.dirname(moc_path), exist_ok=True)
        title = moc_name.replace("MOC-", "", 1).replace("-", " ").strip().title()
        is_default = moc_name == self.obsidian_default_moc
        body = _new_moc_template(moc_name, title, is_default=is_default)
        with open(moc_path, "w", encoding="utf-8") as f:
            f.write(body)
        logger.info("Created MOC: %s", moc_path)
        return moc_path

    def _link_in_moc(self, moc_name: str, wikilink: str, *, section: str) -> None:
        if not self.obsidian_moc_active():
            return
        moc_path = self._ensure_moc_file(moc_name)
        link_line = f"- [[{wikilink}]]"
        heading = f"## {section}"
        try:
            with open(moc_path, "r", encoding="utf-8") as f:
                content = f.read()
            if f"[[{wikilink}]]" in content:
                return
            if heading in content:
                parts = content.split(heading, 1)
                rest = parts[1]
                if "\n## " in rest:
                    head, tail = rest.split("\n## ", 1)
                    rest = head.rstrip() + "\n" + link_line + "\n\n## " + tail
                else:
                    rest = rest.rstrip() + "\n" + link_line + "\n"
                content = parts[0] + heading + rest
            else:
                content = content.rstrip() + f"\n\n{heading}\n\n{link_line}\n"
            with open(moc_path, "w", encoding="utf-8") as f:
                f.write(content)
        except OSError as exc:
            logger.warning("Could not update MOC %s: %s", moc_path, exc)

    def decrypt_text_report(self, path: str) -> str:
        """Decrypt a .txt.enc file and return its contents as a string."""
        if not self._enc_key:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
        with open(path, "rb") as f:
            return self._fernet().decrypt(f.read()).decode("utf-8")

    # ── Rotation ───────────────────────────────────────────────────────────────

    def _rotate_old_reports(self) -> None:
        cutoff = datetime.now(timezone.utc) - timedelta(days=self.retention_days)
        deleted = 0
        for root, _dirs, files in os.walk(self.reports_dir):
            for fname in files:
                fpath = os.path.join(root, fname)
                if not os.path.isfile(fpath):
                    continue
                mtime = datetime.fromtimestamp(
                    os.path.getmtime(fpath), tz=timezone.utc,
                )
                if mtime < cutoff:
                    try:
                        os.remove(fpath)
                        deleted += 1
                    except OSError as exc:
                        logger.warning(
                            "Could not delete old report %s: %s", fpath, exc,
                        )
        if deleted:
            logger.info(
                "Rotated %d report(s) older than %d days",
                deleted, self.retention_days,
            )

    # ── Encryption helpers ─────────────────────────────────────────────────────

    def _fernet(self):
        from cryptography.fernet import Fernet
        return Fernet(_derive_fernet_key(self._enc_key))

    def _encrypt_excel(self, path: str) -> str:
        """Password-protect the xlsx in-place using msoffcrypto-tool."""
        try:
            import msoffcrypto
            encrypted_path = path  # overwrite the unencrypted file
            buf = io.BytesIO()
            with open(path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.encrypt(self._enc_key, buf)
            with open(encrypted_path, "wb") as f:
                f.write(buf.getvalue())
            return encrypted_path
        except Exception as exc:
            logger.warning("Excel encryption failed (file saved unencrypted): %s", exc)
            return path


# ── Helpers ────────────────────────────────────────────────────────────────────

def _env_subdir(env_key: str, default: str) -> str:
    return (os.getenv(env_key, default) or default).strip().strip("/\\")


def _date_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")


def _sanitize_knowledge_slug(raw: str) -> str:
    slug = raw.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = slug.strip("-")
    return slug[:60] if slug else ""


def _new_moc_template(moc_name: str, title: str, *, is_default: bool) -> str:
    intro = (
        "Map of content for portfolio analysis runs and durable notes from the stock analyser."
        if is_default
        else f"Map of content for {title} (linked from stock analyser)."
    )
    related = ""
    if is_default:
        related = "\n## Related\n\n- [[MOC-uk-reits]]\n"
    return f"""---
tags:
  - moc
  - stock-analyzer
---

# {moc_name}

{intro}

## Full Portfolio runs

## Individual Stock runs

## Knowledge notes
{related}"""


def _find_note_by_slug(notes_dir: str, slug: str) -> Optional[str]:
    """Return filename if a note ending with -{slug}.md already exists."""
    suffix = f"-{slug}.md"
    try:
        for fname in os.listdir(notes_dir):
            if fname.endswith(suffix):
                return fname
    except OSError:
        pass
    return None


def _write_header_row(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1A237E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
